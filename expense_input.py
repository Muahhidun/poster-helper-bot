"""
Expense Input Module - обработка расходов из листа кассира и Kaspi выписки

Workflow:
1. Пользователь скидывает фото/текст
2. OCR распознаёт текст
3. GPT парсит в список расходов с типами (транзакция/поставка)
4. Пользователь подтверждает/редактирует
5. Создаются транзакции в Poster
"""

import json
import logging
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from enum import Enum
from datetime import datetime

from google.cloud import documentai_v1 as documentai
from google.oauth2 import service_account
from openai import OpenAI

from config import (
    GOOGLE_CLOUD_PROJECT_ID,
    GOOGLE_CLOUD_LOCATION,
    GOOGLE_DOCAI_OCR_PROCESSOR_ID,
    GOOGLE_APPLICATION_CREDENTIALS_JSON,
    OPENAI_API_KEY
)

logger = logging.getLogger(__name__)

# OpenAI клиент — ленивая инициализация, чтобы импорт модуля
# (например, ради parse_kaspi_xlsx) не требовал OPENAI_API_KEY
_openai_client = None


def _get_openai_client():
    global _openai_client
    if _openai_client is None:
        _openai_client = OpenAI(api_key=OPENAI_API_KEY)
    return _openai_client


class ExpenseType(Enum):
    """Тип расхода"""
    TRANSACTION = "транзакция"  # Простой расход (услуги, зарплаты)
    SUPPLY = "поставка"  # Закуп товаров (нужна накладная)


@dataclass
class ExpenseItem:
    """Одна позиция расхода"""
    amount: float  # Сумма в тенге
    description: str  # Описание/комментарий
    expense_type: ExpenseType  # Тип: транзакция или поставка
    category: Optional[str] = None  # Категория расхода
    source: str = "наличка"  # Источник: наличка, kaspi

    # Для поставок
    quantity: Optional[float] = None
    unit: Optional[str] = None
    price_per_unit: Optional[float] = None

    # Информация о поставщике (из алиасов)
    supplier_id: Optional[int] = None
    supplier_name: Optional[str] = None
    original_beneficiary: Optional[str] = None  # Оригинальное название из Kaspi

    # Идентификатор для кнопок
    id: str = field(default_factory=lambda: "")

    def __post_init__(self):
        if not self.id:
            # Генерируем уникальный ID
            import hashlib
            data = f"{self.amount}{self.description}{datetime.now().timestamp()}"
            self.id = hashlib.md5(data.encode()).hexdigest()[:8]


@dataclass
class ExpenseSession:
    """Сессия ввода расходов"""
    items: List[ExpenseItem] = field(default_factory=list)
    source_account: str = "Оставил в кассе (на закупы)"  # Счёт списания
    created_at: datetime = field(default_factory=datetime.now)

    def get_transactions(self) -> List[ExpenseItem]:
        """Получить только транзакции"""
        return [i for i in self.items if i.expense_type == ExpenseType.TRANSACTION]

    def get_supplies(self) -> List[ExpenseItem]:
        """Получить только поставки"""
        return [i for i in self.items if i.expense_type == ExpenseType.SUPPLY]

    def total_amount(self) -> float:
        """Общая сумма"""
        return sum(i.amount for i in self.items)

    def toggle_type(self, item_id: str) -> bool:
        """Переключить тип расхода (транзакция <-> поставка)"""
        for item in self.items:
            if item.id == item_id:
                if item.expense_type == ExpenseType.TRANSACTION:
                    item.expense_type = ExpenseType.SUPPLY
                else:
                    item.expense_type = ExpenseType.TRANSACTION
                return True
        return False


# Категории расходов и ключевые слова для определения
CATEGORY_KEYWORDS = {
    "Зарплаты": ["зарплата", "зп", "курьер", "кассир", "повар", "оплата труда", "аванс"],
    "Хозтовары": ["мыло", "моющее", "салфетки", "туалетная", "губки", "перчатки", "мусорные", "пакеты"],
    "Транспорт": ["такси", "доставка", "яндекс", "убер", "бензин", "топливо"],
    "Коммуналка": ["свет", "электричество", "вода", "газ", "отопление", "аренда", "интернет"],
    "Ремонт": ["ремонт", "запчасти", "сантехник", "электрик"],
    "Реклама": ["реклама", "баннер", "флаер", "instagram", "smm"],
    "Канцелярия": ["канцелярия", "бумага", "ручки", "скотч", "файлы"],
}

# Ключевые слова для определения поставок (закуп продуктов)
SUPPLY_KEYWORDS = [
    # Мясо
    "фарш", "крыло", "курица", "говядина", "свинина", "мясо", "бедро", "филе",
    # Молочка
    "сыр", "молоко", "сметана", "творог", "масло", "пармезан", "чеддер", "моцарелла",
    # Овощи
    "овощи", "помидор", "огурец", "лук", "картофель", "морковь", "капуста", "перец",
    "кюрдамир", "зелень", "салат",
    # Другие продукты
    "мука", "соус", "кетчуп", "майонез", "колбаса", "сосиски", "яйца",
    # Напитки
    "кола", "спрайт", "вода", "сок", "напиток",
    # Поставщики
    "арзан", "магнум", "метро", "япоша", "идея", "сарыарка",
]


def detect_expense_type(description: str) -> ExpenseType:
    """Определить тип расхода по описанию"""
    desc_lower = description.lower()

    # Проверяем на поставку
    for keyword in SUPPLY_KEYWORDS:
        if keyword in desc_lower:
            return ExpenseType.SUPPLY

    # По умолчанию - транзакция
    return ExpenseType.TRANSACTION


def detect_category(description: str) -> Optional[str]:
    """Определить категорию расхода по описанию"""
    desc_lower = description.lower()

    for category, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            if keyword in desc_lower:
                return category

    return "Прочее"


def get_docai_client():
    """Создать клиент Document AI"""
    if not GOOGLE_APPLICATION_CREDENTIALS_JSON:
        raise ValueError("GOOGLE_APPLICATION_CREDENTIALS_JSON не установлен")

    credentials_dict = json.loads(GOOGLE_APPLICATION_CREDENTIALS_JSON)
    credentials = service_account.Credentials.from_service_account_info(credentials_dict)

    opts = {"api_endpoint": f"{GOOGLE_CLOUD_LOCATION}-documentai.googleapis.com"}
    client = documentai.DocumentProcessorServiceClient(
        credentials=credentials,
        client_options=opts
    )

    return client


async def ocr_image(image_path: str) -> str:
    """Распознать текст с изображения через Document AI"""
    with open(image_path, 'rb') as f:
        image_content = f.read()

    docai_client = get_docai_client()

    processor_name = docai_client.processor_path(
        GOOGLE_CLOUD_PROJECT_ID,
        GOOGLE_CLOUD_LOCATION,
        GOOGLE_DOCAI_OCR_PROCESSOR_ID
    )

    raw_document = documentai.RawDocument(
        content=image_content,
        mime_type="image/jpeg"
    )

    request = documentai.ProcessRequest(
        name=processor_name,
        raw_document=raw_document
    )

    result = docai_client.process_document(request=request)
    return result.document.text


async def parse_handwritten_with_vision(image_path: str, source: str = "наличка") -> List['ExpenseItem']:
    """
    Распознать рукописный лист расходов напрямую через GPT-4 Vision.
    Гораздо лучше справляется с почерком чем Document AI OCR.
    """
    import base64

    logger.info(f"🔍 Распознаю рукописный текст через GPT-4 Vision: {image_path}")

    # Читаем и кодируем изображение
    with open(image_path, 'rb') as f:
        image_data = f.read()
    base64_image = base64.b64encode(image_data).decode('utf-8')

    # Определяем mime type
    if image_path.lower().endswith('.png'):
        mime_type = "image/png"
    else:
        mime_type = "image/jpeg"

    prompt = """Внимательно прочитай этот рукописный лист расходов кассира.

ВАЖНО: Это может быть список зарплат сотрудникам или список закупок.

Формат листа обычно:
- Имя сотрудника / название товара — сумма в тенге
- Может быть в 2 колонки (имена слева, суммы справа)
- Суммы могут быть написаны рядом с именами или под ними

Извлеки ВСЕ позиции. Для каждой определи:
1. amount - сумма в тенге (число)
2. description - описание (имя сотрудника или что купили)
3. type - тип: "транзакция" (зарплаты, услуги) или "поставка" (продукты)

Если видишь имена людей (Бека, Батима, Курьер и т.д.) - это скорее всего зарплаты = "транзакция".
Если видишь продукты (мясо, овощи, сыр) - это "поставка".

Верни JSON:
{
    "items": [
        {"amount": 5000, "description": "Бека", "type": "транзакция"},
        {"amount": 10000, "description": "Курьер", "type": "транзакция"}
    ]
}

ВАЖНО:
- Прочитай ВСЕ записи, даже если почерк неразборчивый
- Если не можешь прочитать имя точно - напиши как понял
- Суммы должны быть точными числами без пробелов"""

    try:
        response = _get_openai_client().chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": prompt
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{base64_image}",
                                "detail": "high"
                            }
                        }
                    ]
                }
            ],
            max_tokens=2000,
            temperature=0.1,
            response_format={"type": "json_object"}
        )

        result_text = response.choices[0].message.content.strip()
        logger.info(f"📄 GPT-4 Vision ответ: {result_text[:500]}")

        data = json.loads(result_text)

        items = []
        for item_data in data.get('items', []):
            expense_type = (
                ExpenseType.SUPPLY
                if item_data.get('type') == 'поставка'
                else ExpenseType.TRANSACTION
            )

            item = ExpenseItem(
                amount=float(item_data.get('amount', 0)),
                description=str(item_data.get('description', '')),
                expense_type=expense_type,
                category=detect_category(str(item_data.get('description', ''))),
                source=source,
                quantity=item_data.get('quantity'),
                unit=item_data.get('unit'),
                price_per_unit=item_data.get('price_per_unit')
            )
            items.append(item)

        logger.info(f"✅ GPT-4 Vision распознал {len(items)} позиций")
        return items

    except Exception as e:
        logger.error(f"Ошибка GPT-4 Vision: {e}")
        raise


async def parse_cashier_sheet(ocr_text: str, source: str = "наличка") -> List[ExpenseItem]:
    """
    Распарсить лист кассира через GPT-4

    Args:
        ocr_text: Распознанный текст с листа
        source: Источник средств (наличка, kaspi)

    Returns:
        Список ExpenseItem
    """
    prompt = f"""
Вот текст листа расходов кассира (распознан через OCR):

---
{ocr_text}
---

Извлеки ВСЕ расходы в JSON формате. Для каждого расхода определи:
1. amount - сумма в тенге (число)
2. description - описание (что купили/оплатили)
3. type - тип: "транзакция" (услуги, зарплаты, хозтовары) или "поставка" (продукты питания для ресторана)

Правила определения типа:
- "поставка" = закуп продуктов: мясо, овощи, молочка, напитки, ингредиенты
- "транзакция" = всё остальное: зарплаты, такси, хозтовары, ремонт, услуги

Примеры:
- "Фарш 12кг 33600" → поставка (мясо)
- "Зарплата курьеру 15000" → транзакция (зарплата)
- "Овощи Кюрдамир 8500" → поставка (овощи)
- "Мыломойка 3500" → транзакция (хозтовары)

Если есть количество и цена за единицу, извлеки их тоже:
- quantity - количество (число)
- unit - единица (кг, шт, л, упак)
- price_per_unit - цена за единицу

Верни JSON:
{{
    "items": [
        {{
            "amount": 33600,
            "description": "Фарш",
            "type": "поставка",
            "quantity": 12,
            "unit": "кг",
            "price_per_unit": 2800
        }},
        {{
            "amount": 15000,
            "description": "Зарплата курьеру",
            "type": "транзакция"
        }}
    ]
}}

ВАЖНО:
- Извлекай ВСЕ строки расходов
- Суммы должны быть точными
- Если не можешь определить тип - ставь "транзакция"
"""

    response = _get_openai_client().chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "Ты система извлечения данных из рукописных листов расходов. Возвращаешь валидный JSON."
            },
            {
                "role": "user",
                "content": prompt
            }
        ],
        max_tokens=3000,
        temperature=0.1,
        response_format={"type": "json_object"}
    )

    result_text = response.choices[0].message.content.strip()
    data = json.loads(result_text)

    items = []
    for item_data in data.get('items', []):
        expense_type = (
            ExpenseType.SUPPLY
            if item_data.get('type') == 'поставка'
            else ExpenseType.TRANSACTION
        )

        item = ExpenseItem(
            amount=float(item_data['amount']),
            description=item_data['description'],
            expense_type=expense_type,
            category=detect_category(item_data['description']),
            source=source,
            quantity=item_data.get('quantity'),
            unit=item_data.get('unit'),
            price_per_unit=item_data.get('price_per_unit')
        )
        items.append(item)

    return items


def detect_source_from_ocr(ocr_text: str) -> str:
    """
    Определить источник платежа по OCR-тексту

    Ищем признаки Kaspi: "Kaspi", "Каспи", "QR", характерные элементы интерфейса.

    Returns:
        "kaspi" или "наличка"
    """
    text_lower = ocr_text.lower()

    # Признаки Kaspi скриншота
    kaspi_indicators = [
        "kaspi",
        "каспи",
        "kaspi.kz",
        "kaspi pay",
        "kaspi gold",
        "qr-код",
        "qr код",
        "перевод",  # часто в скриншотах переводов
        "платёж успешен",
        "платеж успешен",
        "получатель",  # характерно для скриншота перевода
        "отправитель",
        "комиссия",
    ]

    kaspi_count = sum(1 for indicator in kaspi_indicators if indicator in text_lower)

    # Если 2+ признака Kaspi - считаем что это Kaspi скриншот
    if kaspi_count >= 2:
        logger.info(f"📱 Определён источник Kaspi ({kaspi_count} признаков)")
        return "kaspi"

    # Если есть прямое упоминание Kaspi
    if "kaspi" in text_lower or "каспи" in text_lower:
        logger.info("📱 Определён источник Kaspi (по названию)")
        return "kaspi"

    logger.info("💵 Определён источник: наличка")
    return "наличка"


async def parse_cashier_sheet_from_image(image_path: str, source: str = None, use_vision: bool = True) -> List[ExpenseItem]:
    """
    Распознать и распарсить лист кассира с фото

    Args:
        image_path: Путь к фото
        source: Источник средств (если None - определяется автоматически)
        use_vision: Использовать GPT-4 Vision (лучше для рукописного текста)

    Returns:
        Список ExpenseItem
    """
    logger.info(f"🔍 Распознаю лист кассира: {image_path}")

    # Сначала делаем OCR для определения источника
    try:
        ocr_text = await ocr_image(image_path)
        logger.info(f"📄 OCR получен: {len(ocr_text)} символов")

        # Автоопределение источника
        if source is None:
            source = detect_source_from_ocr(ocr_text)
    except Exception as e:
        logger.warning(f"OCR не удался: {e}, используем source по умолчанию")
        ocr_text = ""
        if source is None:
            source = "наличка"

    # Используем GPT-4 Vision для лучшего распознавания рукописного текста
    if use_vision:
        try:
            items = await parse_handwritten_with_vision(image_path, source)
            if items:
                logger.info(f"✅ Vision распознал {len(items)} позиций (источник: {source})")
                return items
        except Exception as e:
            logger.warning(f"Vision не удался: {e}, пробуем через OCR+GPT")

    # Fallback: традиционный OCR + GPT парсинг
    if ocr_text:
        items = await parse_cashier_sheet(ocr_text, source)
        logger.info(f"✅ OCR+GPT распознал {len(items)} позиций (источник: {source})")
        return items

    return []


async def parse_cashier_sheet_from_url(image_url: str, source: str = None) -> List[ExpenseItem]:
    """Распознать лист кассира по URL (source определяется автоматически если не указан)"""
    import aiohttp
    import tempfile
    import os

    async with aiohttp.ClientSession() as session:
        async with session.get(image_url) as response:
            if response.status != 200:
                raise Exception(f"Не удалось скачать: HTTP {response.status}")
            image_data = await response.read()

    with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as tmp_file:
        tmp_file.write(image_data)
        tmp_path = tmp_file.name

    try:
        return await parse_cashier_sheet_from_image(tmp_path, source)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def format_expense_list(session: ExpenseSession) -> str:
    """Форматировать список расходов для отображения в боте"""
    lines = [f"📋 **{session.source_account}**\n"]

    for i, item in enumerate(session.items, 1):
        type_emoji = "📦" if item.expense_type == ExpenseType.SUPPLY else "💰"
        type_label = "поставка" if item.expense_type == ExpenseType.SUPPLY else "транзакция"

        lines.append(f"{i}. {type_emoji} {item.amount:,.0f}₸ ({type_label})")
        lines.append(f"   └ {item.description}")
        if item.category and item.expense_type == ExpenseType.TRANSACTION:
            lines.append(f"   └ Категория: {item.category}")
        lines.append("")

    # Итоги
    transactions = session.get_transactions()
    supplies = session.get_supplies()

    lines.append("─" * 25)
    lines.append(f"💰 Транзакций: {len(transactions)} на {sum(t.amount for t in transactions):,.0f}₸")
    lines.append(f"📦 Поставок: {len(supplies)} на {sum(s.amount for s in supplies):,.0f}₸")
    lines.append(f"**Итого: {session.total_amount():,.0f}₸**")

    return "\n".join(lines)


async def create_transactions_in_poster(
    session: ExpenseSession,
    telegram_user_id: int,
    account_id: int,
    category_map: Dict[str, int]
) -> Tuple[int, int, List[str]]:
    """
    Создать транзакции в Poster

    Args:
        session: Сессия с расходами
        telegram_user_id: ID пользователя Telegram
        account_id: ID счёта в Poster
        category_map: Маппинг название категории -> ID в Poster

    Returns:
        (успешно, ошибок, список ошибок)
    """
    from database import get_database
    from poster_client import PosterClient

    db = get_database()
    accounts = db.get_accounts(telegram_user_id)

    if not accounts:
        return 0, 0, ["Нет подключенных аккаунтов Poster"]

    # Берём первый аккаунт
    account = accounts[0]

    client = PosterClient(
        telegram_user_id=telegram_user_id,
        poster_token=account['poster_token'],
        poster_user_id=account['poster_user_id'],
        poster_base_url=account['poster_base_url']
    )

    success_count = 0
    error_count = 0
    errors = []

    try:
        transactions = session.get_transactions()

        for item in transactions:
            try:
                # Определяем категорию
                category_id = category_map.get(item.category, category_map.get("Прочее", 1))

                # Создаём транзакцию
                await client.create_transaction(
                    transaction_type=0,  # expense
                    category_id=category_id,
                    account_from_id=account_id,
                    amount=int(item.amount),
                    comment=item.description
                )

                success_count += 1
                logger.info(f"✅ Создана транзакция: {item.amount}₸ - {item.description}")

            except Exception as e:
                error_count += 1
                errors.append(f"{item.description}: {str(e)}")
                logger.error(f"❌ Ошибка создания транзакции: {e}")

    finally:
        await client.close()

    return success_count, error_count, errors


def extract_supplier_name_from_purpose(purpose: str) -> str:
    """
    Извлечь имя поставщика из назначения платежа.

    Примеры:
    - "ИП ЕРЖАНОВА. Оплата с Kaspi QR" → "ИП ЕРЖАНОВА"
    - "ИнариН. Оплата с Kaspi QR" → "ИнариН"
    - "Yaposha Market. Оплата с Kaspi QR" → "Yaposha Market"
    - "Мясной магазин. Оплата" → "Мясной магазин"
    - "Перевод собственных средств на карту Kaspi" → ""
    """
    if not purpose:
        return ""

    purpose = purpose.strip()

    # Паттерны которые означают "не поставщик"
    skip_patterns = [
        "перевод собственных",
        "за профессиональные",
        "за научные",
        "за технические",
    ]
    purpose_lower = purpose.lower()
    for pattern in skip_patterns:
        if pattern in purpose_lower:
            return ""

    # Ищем паттерн "Имя. Оплата" или "Имя. Что-то"
    # Разделитель - точка с пробелом или просто точка
    if ". " in purpose:
        name = purpose.split(". ")[0].strip()
        return name
    elif "." in purpose and not purpose.endswith("."):
        # Если есть точка, но нет пробела после - пробуем разделить
        parts = purpose.split(".")
        if len(parts) >= 2 and len(parts[0]) > 2:
            return parts[0].strip()

    return ""


def parse_kaspi_xlsx(file_path: str, telegram_user_id: int = None) -> List[ExpenseItem]:
    """
    Парсинг выписки Kaspi из XLSX файла

    Колонки:
    - № документа (A/1)
    - Дата операции (B/2)
    - Дебет (C/3) - расход
    - Кредит (D/4) - приход
    - Наименование бенефициара (E/5)
    - ИИК бенефициара (F/6)
    - БИК банка (G/7)
    - КНП (H/8)
    - Назначение платежа (I/9) - ТУТ ИМЕНА ПОСТАВЩИКОВ!

    Args:
        file_path: Путь к XLSX файлу
        telegram_user_id: ID пользователя для поиска алиасов поставщиков

    Returns:
        Список ExpenseItem (только расходы)
    """
    from openpyxl import load_workbook

    logger.info(f"📄 Парсинг Kaspi выписки: {file_path}")

    # Загружаем функцию поиска алиасов поставщиков
    supplier_aliases_lookup = None
    if telegram_user_id:
        try:
            from database import get_database
            db = get_database()
            supplier_aliases_lookup = db.get_supplier_by_alias
        except Exception as e:
            logger.warning(f"Не удалось загрузить алиасы поставщиков: {e}")

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    items = []
    header_found = False
    data_start_row = 0

    # Ищем заголовок таблицы (строка с "№ документа" или "Дата операции")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row and any(cell and "документа" in str(cell).lower() for cell in row if cell):
            header_found = True
            data_start_row = row_idx + 2  # Пропускаем заголовок и номера колонок
            break

    if not header_found:
        # Пробуем найти первую строку с данными
        data_start_row = 1

    # Читаем данные
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row or len(row) < 9:
            continue

        # Пропускаем итоговые строки
        first_cell = str(row[0] or "").lower()
        if "итого" in first_cell or not row[0]:
            continue

        try:
            # Колонки: № док, Дата, Дебет, Кредит, Наименование, ИИК, БИК, КНП, Назначение
            doc_num = row[0]
            date_cell = row[1]
            debit = row[2]  # Расход
            credit = row[3]  # Приход
            beneficiary = row[4]  # Наименование бенефициара / отправителя
            purpose = row[8]  # Назначение платежа (здесь имена поставщиков!)

            # Пропускаем если нет дебета (не расход)
            if not debit or debit == 0:
                continue

            # Парсим сумму
            if isinstance(debit, str):
                amount = float(debit.replace(" ", "").replace(",", "."))
            else:
                amount = float(debit)

            if amount <= 0:
                continue

            # Извлекаем данные
            beneficiary_str = str(beneficiary or "").strip()
            purpose_str = str(purpose or "").strip()

            # Убираем технические данные из beneficiary
            beneficiary_clean = beneficiary_str
            beneficiary_clean = beneficiary_clean.replace("АО \"KASPI BANK\"", "").strip()
            beneficiary_clean = beneficiary_clean.replace("ИИН/БИН", "").strip()
            # Убираем ИИН/БИН номера (971240001315 и т.д.)
            beneficiary_clean = re.sub(r'\d{12}', '', beneficiary_clean).strip()
            beneficiary_clean = beneficiary_clean.strip()

            # 1. Извлекаем имя поставщика из "Назначение платежа"
            #    Например: "ИП ЕРЖАНОВА. Оплата с Kaspi QR" → "ИП ЕРЖАНОВА"
            supplier_from_purpose = extract_supplier_name_from_purpose(purpose_str)

            # 2. Определяем что использовать как имя поставщика для поиска алиаса
            #    Приоритет: purpose > beneficiary
            search_name = supplier_from_purpose if supplier_from_purpose else beneficiary_clean

            # 3. Ищем поставщика по алиасу
            supplier_id = None
            supplier_name = None
            matched_alias = None

            if supplier_aliases_lookup and search_name:
                # Пробуем найти по полному имени
                supplier_match = supplier_aliases_lookup(telegram_user_id, search_name)
                if supplier_match:
                    supplier_id = supplier_match['poster_supplier_id']
                    supplier_name = supplier_match['poster_supplier_name']
                    matched_alias = search_name
                    logger.info(f"🏪 Найден поставщик: '{search_name}' → {supplier_name}")

                # Если не нашли и есть beneficiary - пробуем по нему
                if not supplier_id and beneficiary_clean and beneficiary_clean != search_name:
                    supplier_match = supplier_aliases_lookup(telegram_user_id, beneficiary_clean)
                    if supplier_match:
                        supplier_id = supplier_match['poster_supplier_id']
                        supplier_name = supplier_match['poster_supplier_name']
                        matched_alias = beneficiary_clean
                        logger.info(f"🏪 Найден поставщик по бенефициару: '{beneficiary_clean}' → {supplier_name}")

            # 4. Формируем description
            #    Если нашли поставщика - используем его имя из Poster
            #    Иначе - используем извлечённое имя или beneficiary
            if supplier_name:
                description = supplier_name
            elif supplier_from_purpose:
                description = supplier_from_purpose
            elif beneficiary_clean:
                description = beneficiary_clean
            else:
                description = purpose_str[:50] if purpose_str else f"Расход {doc_num}"

            # 5. Определяем тип
            #    Если нашли поставщика по алиасу - это поставка
            if supplier_id:
                expense_type = ExpenseType.SUPPLY
            else:
                expense_type = detect_expense_type(description)

            category = detect_category(description)

            item = ExpenseItem(
                amount=amount,
                description=description[:100],  # Ограничиваем длину
                expense_type=expense_type,
                category=category,
                source="kaspi",
                supplier_id=supplier_id,
                supplier_name=supplier_name,
                original_beneficiary=beneficiary_str[:100] if beneficiary_str else None
            )
            items.append(item)

            logger.debug(f"Добавлен расход: {amount}₸ - {description[:50]}")

        except (ValueError, TypeError) as e:
            logger.debug(f"Пропуск строки: {e}")
            continue

    wb.close()

    logger.info(f"✅ Распознано {len(items)} расходов из Kaspi выписки")
    return items


async def parse_kaspi_xlsx_from_file(file_path: str) -> List[ExpenseItem]:
    """Асинхронная обёртка для parse_kaspi_xlsx"""
    import asyncio
    return await asyncio.get_event_loop().run_in_executor(None, parse_kaspi_xlsx, file_path)
